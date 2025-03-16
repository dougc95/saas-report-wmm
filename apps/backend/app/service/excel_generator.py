import openpyxl
from app.models.wmm_model import WMMModel
from io import BytesIO


class ExcelGenerator:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "WMM"
        self.row_index = 1

    def add_inputs_sheet(self, lat: float, lon: float, altitude: float, altitude_unit: str, start_date: str, end_date: str, step_days: int):
        inputs_sheet = self.workbook.create_sheet("Inputs")
        inputs_sheet.append(["Parameter", "Value"])
        inputs_sheet.append(["Latitude", lat])
        inputs_sheet.append(["Longitude", lon])
        inputs_sheet.append(["Altitude", altitude])
        inputs_sheet.append(["Altitude Unit", altitude_unit])
        inputs_sheet.append(["Start Date", start_date])
        inputs_sheet.append(["End Date", end_date])
        inputs_sheet.append(["Step (days)", step_days])

    def generate_header(self):
        headers = [
            "Date",
            "Total Field",
            "Horizontal",
            "North",
            "East",
            "Vertical",
            "Declination",
            "Inclination",
            "",
            "Δ Total Field",
            "Δ Horizontal",
            "Δ North",
            "Δ East",
            "Δ Vertical",
            "Δ Declination",
            "Δ Inclination",
        ]
        for col_index, header in enumerate(headers, start=1):
            self.sheet.cell(row=self.row_index, column=col_index, value=header)
        self.row_index += 1

    def add_data(self, model: WMMModel, variation: dict):
        data = [
            model.date,
            model.ti,
            model.bh,
            model.bx,
            model.by,
            model.bz,
            model.dec,
            model.dip,
            "",
            variation["ti"],
            variation["bh"],
            variation["bx"],
            variation["by"],
            variation["bz"],
            variation["dec"],
            variation["dip"],
        ]
        for col_index, value in enumerate(data, start=1):
            self.sheet.cell(row=self.row_index, column=col_index, value=value)
        self.row_index += 1

    def get_buffer(self) -> BytesIO:
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
